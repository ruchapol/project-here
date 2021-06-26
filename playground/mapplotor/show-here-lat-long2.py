from __future__ import unicode_literals

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import openpyxl
import mpu
import random
import re
from graph_imp import Graph

def to_float(number):
    try:
        return float(number)
    except (ValueError, TypeError):
        return -1.0

def to_lat_long(str):

    # if(len(splited_str) == 2):
    try:
        splited_str = str.split(',')
        return (float(splited_str[0]), float(splited_str[1]))
    except (ValueError, TypeError, AttributeError, IndexError):
        return None
    # else:
        # return None

def to_str(data):
    return str(data) if (data is not None) else None

class RoadSegment:

    seg_dict = {}
    road_dict = {}
    road_col_dict = {}
    def __init__(self, li, pc, name, roadName, lat, long):
        self.li = li
        self.pc = pc
        self.name = name
        self.roadName = roadName
        self.lat = lat
        self.long = long
        self.seg_dict[(li, pc)] = self
        self.next_seg = []
        self.prev_seg = []
        try:
            self.road_dict[li].add(self)
        except KeyError:
            self.road_dict[li] = set([self])
            self.road_col_dict[li] = random.choice(['g','r','b','c','y','k'])

# ,'m'


    def __repr__(self):
        return "%s %s %s" % (self.li, self.pc, self.name)

    def __str__(self):
        return "%s %s %s" % (self.li, self.pc, self.name)

    def assign_next(self, li, pc):
        # self.seg_dict[(li, pc)]
        # print(self.seg_dict[(li, pc)])
        if((li, pc) not in self.seg_dict):
            assert type(li) == str
            assert type(pc) == str or pc == None
            # self.seg_dict[(li, pc)]
            raise(KeyError("No pc <%s %s> in pc dict" % (li, pc, )))
        friend_seg = self.seg_dict[(li, pc)]
        self.next_seg.append(friend_seg)
        friend_seg.prev_seg.append(self)

    def assign_prev(self, li, pc):

        # self.seg_dict[(li, pc)]
        # print(self.seg_dict[(li, pc)])
        if((li, pc) not in self.seg_dict):
            assert type(li) == str
            assert type(pc) == str or pc == None
            # self.seg_dict[(li, pc)]
            raise(KeyError("No pc <%s %s> in pc dict" % (li, pc, )))

        friend_seg = self.seg_dict[(li, pc)]
        self.prev_seg.append(friend_seg)
        friend_seg.next_seg.append(self)

    def get_road_color(li):
        return RoadSegment.road_col_dict[li]

    def get_road_list():
        return [road_code for road_code in RoadSegment.road_dict]

    def get_segmemnt_road_set(li):
        return RoadSegment.road_dict[li]

    def get_adjacant_seg(self):
        return self.prev_seg + self.next_seg

    def get_seg_road_order(li):
        mid_segment = next(iter(RoadSegment.road_dict[li]))
        segment = mid_segment
        next_seg = []
        prev_seg = []

        while(len(segment.next_seg) != 0):
            found_same_road = False
            for n_seg in segment.next_seg:
                if(n_seg.li == li):
                    next_seg.append(n_seg)
                    segment = n_seg
                    found_same_road = True
                    break
            if(not found_same_road):
                break

        segment = mid_segment

        while(len(segment.prev_seg) != 0):
            found_same_road = False
            for n_seg in segment.prev_seg:
                if(n_seg.li == li):
                    prev_seg.append(n_seg)
                    segment = n_seg
                    found_same_road = True
                    break
            if(not found_same_road):
                break

        return prev_seg + next_seg

    def get_next_seg(self):
        return self.next_seg

    def get_seg_dict():
        return RoadSegment.seg_dict

    def get_road_name(self):
        return self.roadName

    def get_seg_name(self):
        return self.name

    def get_long_lat(self):
        return self.long, self.lat

    def get_road_long_lat(li):
        lats = [roadSeg.lat for roadSeg in RoadSegment.road_dict[li]]
        longs = [roadSeg.long for roadSeg in RoadSegment.road_dict[li]]
        return np.average(longs), np.average(lats)

    def get_road_name_by_li(li):
        seg_tmp = next(iter(RoadSegment.road_dict[li]))
        return seg_tmp.get_road_name()


# font_list = fm.createFontList(['THSarabunNew.ttf'])
# font_list = fm.
plt.rcParams.update({'font.size': 10})
# fm.fontManager.ttflist.extend(font_list)
# set font
plt.rcParams['font.family'] = 'TH Sarabun New'

# geo_n_book = openpyxl.load_workbook(r"C:\Users\ak74b\Desktop\work\4_1\sP\python\routing\geo_with_ajacant.xlsx")
geo_n_book = openpyxl.load_workbook(r"C:\Users\ak74b\Desktop\work\4_1\sP\traffic_sp\python\routing\geo_with_ajacant.xlsx")

# geo_n_book = openpyxl.load_workbook(r"C:\Users\ak74b\Desktop\work\4_1\sP\python\routing\geo_n2.xlsx")
s_book  = openpyxl.load_workbook(r"C:\Users\ak74b\Desktop\work\4_1\sP\thai-lat-long\Thailand_T19_v3.2_flat_Thai.xlsx")

s_sheet = s_book.active
geo_n_sheet = geo_n_book.active

# lats = np.array([to_float(acell.value) for acell in geo_n_sheet['Q']])
lats_longs = np.array([to_lat_long(acell.value) for acell in geo_n_sheet['E']])
names = np.array([acell.value for acell in geo_n_sheet['D']])
pcs = np.array([acell.value for acell in geo_n_sheet['B']], dtype=object)
li_codes = np.array([acell.value for acell in geo_n_sheet['A']])
road_names = np.array([acell.value for acell in geo_n_sheet['C']])
selected_segment = np.array([acell.value for acell in geo_n_sheet['H']])
next_segment = np.array([[cell.value for cell in row] for row in geo_n_sheet['H':'M']])


pcs_s = np.array([acell.value for acell in s_sheet['E']], dtype=object)
next_pcs_s = np.array([acell.value for acell in s_sheet['O']], dtype=object)
prev_pcs_s = np.array([acell.value for acell in s_sheet['N']], dtype=object)



selected_arr = lats_longs != None

names = names[selected_arr]
pcs = pcs[selected_arr]
lats = [x[0] for x in lats_longs[selected_arr]]
longs = [x[1] for x in lats_longs[selected_arr]]
road_names = road_names[selected_arr]
selected_segment = selected_segment[selected_arr]
li_codes = li_codes[selected_arr]
# print(next_segment.T[0])
# print(next_segment.T[1])
next_segment = next_segment.T[selected_arr, :]

# road_dict = {}
# pc_pos_dict = {pcs[i]: (lats[i], longs[i]) for i in range(len(pcs))}  # pc: (lat, long)
connect_seg_dict = {to_str(pcs_s[i]):(to_str(prev_pcs_s[i]), to_str(next_pcs_s[i])) for i in range(len(pcs_s))}    # {pc:(prev_, next_)}
# seleced_seg_dict = {pcs[i]:selected_segment[i] for i in range(len(pcs))}
# connect_road_dict = {(li_codes, to_str(pcs[i])):(to_str(), to_str())  }

connect_seg_dict["57499"] = ("57498", None)
connect_seg_dict["57498"] = (None, None)
connect_seg_dict["57508"] = (None, None)
connect_seg_dict["57507"] = ("57506", None)
connect_seg_dict["57506"] = (None, "57507")
connect_seg_dict["57504"] = ("57503", None)
connect_seg_dict["57503"] = ("57502", "57504")
connect_seg_dict["57502"] = ("57501", "57503")
connect_seg_dict["57501"] = (None, "57502")


# 219-57496	57499	219-57496	57498
# 219-57496	57498	219-57496
# 219-57505	57507	219-57505	57506
# 219-57505	57506	219-57505
# 219+57505	57506	219+57505	57507
# 219+57505	57507	219+57505
# 219+57496	57499	219+57496
# 219-57500	57504	219-57500	57503
# 219-57500	57503	219-57500	57502
# 219-57500	57502	219-57500	57501
# 219-57500	57501	219-57500
# 219+57500	57501	219+57500	57502
# 219+57500	57502	219+57500	57503
# 219+57500	57503	219+57500	57504
# 219+57500	57504	219+57500

in_count = 0
out_count = 0
for i in range(len(pcs)):
    # self, li, pc, name, roadName, lat, long
    z = re.match(r".*(out).*", str(selected_segment[i]))
    if(z is not None):
        print("pc:", pcs[i], "name:", names[i], "is %s!!" % (selected_segment[i], ))
        out_count += 1
        continue
    RoadSegment(to_str(li_codes[i]), to_str(pcs[i]), names[i], road_names[i], lats[i], longs[i])
    seg_tmp = RoadSegment.get_seg_dict()[(li_codes[i], to_str(pcs[i]))]
    in_count += 1
    print("pc:", pcs[i], "name:", seg_tmp.get_seg_name(), "is %s" % (selected_segment[i], ))

print("IN", in_count, "OUT", out_count, "total", in_count + out_count)

seg_dict = RoadSegment.get_seg_dict()


# segment's connection in same road
for li, pc in seg_dict:
    z = re.match(r"^\d+([-+])\d+$", li)
    if(z.groups()[0] == "+"):
        prev_pcs, next_pcs = connect_seg_dict[pc]
    elif(z.groups()[0] == "-"):
        next_pcs, prev_pcs = connect_seg_dict[pc]

    try:
        seg_dict[(li, pc)].assign_prev(li, prev_pcs)
        # seg_dict[(li, pc)]
        # seg_dict[(li, prev_pcs)]
    except KeyError as ke:
        # print("no (%s, %s) in db" % (li, prev_pcs, ))
        print(str(ke))
    try:
        seg_dict[(li, pc)].assign_next(li, next_pcs)
        # seg_dict[(li, next_pcs)]
        # seg_dict[(li, pc)]
    except KeyError as ke:
        print(str(ke))
        # print("no (%s, %s) in db" % (li, next_pcs, ))
    del prev_pcs
    del next_pcs

for key in seg_dict:
    assert type(key[0]) == str
    assert type(key[1]) == str or pc == None
    tmp_seg = seg_dict[key]
    print(str(tmp_seg), tmp_seg.get_adjacant_seg())

seg_dict = RoadSegment.get_seg_dict()

for i in range(len(pcs)):
    for j in range(0, len(next_segment[i]), 2):
        if(next_segment[i][j] is None or next_segment[i][j+1] is None):
            continue
        try:
            print((li_codes[i], to_str(pcs[i])), next_segment[i][j], next_segment[i][j+1])
            seg_dict[(li_codes[i], to_str(pcs[i]))].assign_next(next_segment[i][j], to_str(next_segment[i][j+1]))
        except KeyError as ke:
            print(ke, '...')
    # connect_road_dict = {(li_codes, to_str(pcs[i])):(to_str(), to_str())  }

fig, ax = plt.subplots()

plt.xlabel("longitude")
plt.ylabel("latitude")

# plot road line by road Group
for li in RoadSegment.get_road_list():
    # seg_order = RoadSegment.get_seg_road_order(li)
    # longs_lats = [road_seg.get_long_lat() for road_seg in seg_order]

    z = re.match(r"^\d+([-+])\d+$", li)
    if(z.groups()[0] == "+"):
        offset = 1e-4
    elif(z.groups()[0] == "-"):
        offset = -1e-4

    seg_set = RoadSegment.get_segmemnt_road_set(li)
    for segment in seg_set:
        long_lat = segment.get_long_lat()
        adjacant_segs = segment.get_adjacant_seg()

        longs = [adj_seg.get_long_lat()[0] for adj_seg in adjacant_segs]
        lats = [adj_seg.get_long_lat()[1] for adj_seg in adjacant_segs]

        main_long = long_lat[0]
        main_lat = long_lat[1]
        # longs += long_lat[0]
        # lats += long_lat[1]
        # get_long_lat
        # get_seg_name

        plt.plot([main_long+offset], [main_lat+offset], marker='o',
                 color=RoadSegment.get_road_color(li)) #1 plot graph


        # plot road line
        for i in range(len(longs)):
            plt.plot([main_long+offset, longs[i]+offset],
                     [main_lat+offset, lats[i]+offset],
                     color=RoadSegment.get_road_color(li)) #2 plot line
    # ax.annotate
    print("li:", li, "group:", ", ".join([str(seg) for seg in seg_set]))
    # show road name and number

    road_name = RoadSegment.get_road_name_by_li(li)
    annt_posx, annt_posy = RoadSegment.get_road_long_lat(li)

    ax.annotate("\n".join([str(x) for x in [li, road_name]]) ,
                (annt_posx, annt_posy), color="red")

# plot road connector



# ,'m'



# plot road names
seg_dict = RoadSegment.get_seg_dict()
for pc in seg_dict:
    # if(selected_segment[i] == 'out'):
    segment = seg_dict[pc]
    #     print("Skipped name:", names[i])
    #     continue
    name = segment.get_seg_name()
    long, lat = segment.get_long_lat()

    ax.annotate("%s\n%s"%(name, pc), (long + random.uniform(-1e-3, 1e-3),
                                      lat + random.uniform(-1e-3, 1e-3)))

# plot destination and source
marker_n = ["พระจอม", "วงศ์สว่าง", "กระทรวง", "บางโพ"]
marker_lat = [13.818851, 13.82972, 13.84858, 13.806172]
marker_long = [100.5138, 100.5266, 100.5147, 100.5215]

plt.scatter(marker_long, marker_lat)

for i in range(len(marker_n)):

    ax.annotate(marker_n[i], (marker_long[i], marker_lat[i]))

seg_dict = RoadSegment.get_seg_dict()
for li, pc in seg_dict:
    tmp_seg = seg_dict[li, pc]
    name = tmp_seg.get_seg_name()
    road_name = tmp_seg.get_road_name()
    node_tmp = Graph(name=(tmp_seg.name), id=(li, pc), cost=1)

for id in Graph.node_dict:
    tmp_seg = seg_dict[id]
    for next_seg_tmp in tmp_seg.get_next_seg():

        tmp_next_node = Graph.node_dict[(next_seg_tmp.li, next_seg_tmp.pc)]
        tmp_this_node = Graph.node_dict[id]
        tmp_this_node.assign_outbound(tmp_next_node)

for li, pc in Graph.node_dict:
    node_tmp = Graph.node_dict[(li, pc)]
    print(node_tmp.name, li, pc)
    for next_node in node_tmp.outbound:
        print('\t-', next_node.name, next_node.id)

Graph.gen_dijk_table(("219-02600", "46772"))
cost_wongsawang_mrt = Graph.travel(Graph.node_dict["219-02960", "47759"])
print("step:", cost_wongsawang_mrt)
cost_bangpho = Graph.travel(Graph.node_dict["219-00661", "35514"])
print("step:", cost_bangpho)
Graph.gen_dijk_table(("219+02600", "46771"))
cost_health_minist = Graph.travel(Graph.node_dict["219+57505", "57507"])
print("step:", cost_health_minist)




# geo_n_book.save(r'C:\Users\ak74b\Desktop\work\4_1\sP\python\routing\geo_n3.xlsx')

plt.show()
